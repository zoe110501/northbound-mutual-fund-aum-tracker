from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_MAP = {
    "sequence": "序号",
    "fund_code": "基金代码",
    "name": "基金全称",
    "inception_date": "基金成立日",
    "recognition_approval_date": "互认基金批复日期",
    "investment_type": "投资类型",
    "investment_region": "投资区域",
    "recognition_status": "互认状态",
    "aum_cny_100m_from_attachment": "基金规模合计(亿人民币)",
    "fund_manager": "基金经理",
    "management_company": "基金管理人",
    "trustee": "基金托管人",
    "mainland_distributors": "代销机构",
    "management_fee_pct": "管理费率(%)",
    "management_fee_note": "管理费说明",
    "trustee_fee_pct": "托管费率(%)",
    "sales_service_fee_pct": "销售服务费(%)",
    "investment_objective": "投资目标",
    "domicile": "注册地",
}

FIELD_ORDER = list(HEADER_MAP)
AUM_FIELD = "aum_cny_100m_from_attachment"
RED_FONT = Font(name="Arial", color="FF0000")
DEFAULT_FONT = Font(name="Arial", color="000000")
HEADER_FONT = Font(name="Arial", bold=True, color="000000")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")


def write_excel_report(payload: dict[str, Any], fund_data_path: Path, output_path: Path, report_date: str | None = None) -> None:
    fund_data = json.loads(fund_data_path.read_text(encoding="utf-8"))
    report_date = report_date or _report_date(payload)
    workbook = Workbook()
    workbook.remove(workbook.active)

    evidence_by_name = {
        "funds": _evidence_by_target(payload, "global_evidence"),
        "mainland_share_classes": _evidence_by_target(payload, "mainland_evidence"),
    }
    errors_by_manager = _errors_by_manager(payload)

    _write_sheet(
        workbook=workbook,
        title="合并成基金全称",
        rows=fund_data["sheets"]["funds"],
        latest_values=evidence_by_name["funds"],
        errors_by_manager=errors_by_manager,
        report_date=report_date,
        name_header="基金全称",
    )
    _write_sheet(
        workbook=workbook,
        title="分各个份额（仅内地销售份额，不含所有份额）",
        rows=fund_data["sheets"]["mainland_share_classes"],
        latest_values=evidence_by_name["mainland_share_classes"],
        errors_by_manager=errors_by_manager,
        report_date=report_date,
        name_header="基金简称",
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _write_sheet(
    *,
    workbook: Workbook,
    title: str,
    rows: list[dict[str, Any]],
    latest_values: dict[str, dict[str, Any]],
    errors_by_manager: dict[str, list[str]],
    report_date: str,
    name_header: str,
) -> None:
    sheet = workbook.create_sheet(title)
    headers = [HEADER_MAP[field] for field in FIELD_ORDER]
    headers[FIELD_ORDER.index("name")] = name_header
    latest_header = f"最新基金规模合计(亿人民币）（{report_date}）"
    insert_at = FIELD_ORDER.index(AUM_FIELD) + 2
    headers.insert(insert_at - 1, latest_header)
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, item in enumerate(rows, start=2):
        values = [item.get(field) for field in FIELD_ORDER]
        target_name = str(item.get("name") or "").strip()
        manager = str(item.get("management_company") or "").strip()
        latest = latest_values.get(target_name)
        latest_value = latest.get("aum_cny_100m") if latest else None
        values.insert(insert_at - 1, latest_value)
        sheet.append(values)

        latest_cell = sheet.cell(row=row_index, column=insert_at)
        latest_cell.number_format = "0.000000"
        original_value = _as_float(item.get(AUM_FIELD))
        if latest_value is not None and original_value is not None and round(float(latest_value) - original_value, 6) != 0:
            latest_cell.font = RED_FONT
        else:
            latest_cell.font = DEFAULT_FONT
        latest_cell.comment = Comment(_cell_note(latest, manager, errors_by_manager), "northbound-aum-tracker")

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if cell.font == RED_FONT:
                continue
            cell.font = DEFAULT_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for index, width in enumerate(_column_widths(len(headers)), start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _evidence_by_target(payload: dict[str, Any], evidence_key: str) -> dict[str, dict[str, Any]]:
    cny_rate = _as_float((payload.get("fx_rates_to_usd") or {}).get("CNY"))
    latest: dict[str, dict[str, Any]] = {}
    for result in payload.get("results", []):
        for evidence in result.get(evidence_key, []):
            target_name = str(evidence.get("target_name") or "").strip()
            if not target_name or target_name in latest:
                continue
            value = _evidence_to_cny_100m(evidence, cny_rate)
            if value is None:
                continue
            latest[target_name] = {
                "aum_cny_100m": round(value, 6),
                "source_url": evidence.get("source_url") or "",
                "currency": evidence.get("currency") or "",
                "context": evidence.get("context") or "",
            }
    return latest


def _evidence_to_cny_100m(evidence: dict[str, Any], cny_rate_to_usd: float | None) -> float | None:
    currency = str(evidence.get("currency") or "").upper()
    amount = _as_float(evidence.get("amount"))
    if currency == "CNY" and amount is not None:
        return amount / 100_000_000
    amount_usd = _as_float(evidence.get("amount_usd"))
    if amount_usd is None or not cny_rate_to_usd:
        return None
    return amount_usd / cny_rate_to_usd / 100_000_000


def _errors_by_manager(payload: dict[str, Any]) -> dict[str, list[str]]:
    return {
        str(result.get("manager") or ""): [str(error) for error in result.get("errors", [])]
        for result in payload.get("results", [])
    }


def _cell_note(latest: dict[str, Any] | None, manager: str, errors_by_manager: dict[str, list[str]]) -> str:
    if latest:
        lines = [
            "来源: 官网/财务报表自动抓取",
            f"URL: {latest.get('source_url', '')}",
            f"原币种: {latest.get('currency', '')}",
        ]
        context = str(latest.get("context") or "").strip()
        if context:
            lines.append(f"上下文: {context[:500]}")
        return "\n".join(lines)

    errors = errors_by_manager.get(manager) or []
    lines = ["未抓到最新规模，单元格留空。"]
    if manager:
        lines.append(f"基金管理人: {manager}")
    if errors:
        lines.append("抓取信息: " + " | ".join(errors[:3])[:700])
    return "\n".join(lines)


def _report_date(payload: dict[str, Any]) -> str:
    run_at = str(payload.get("run_at_utc") or "")
    if run_at:
        return run_at[:10]
    return datetime.now().date().isoformat()


def _as_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _column_widths(column_count: int) -> list[int]:
    base = [8, 14, 34, 12, 16, 12, 12, 12, 16, 22, 18, 20, 20, 36, 12, 24, 12, 14, 48, 12]
    if column_count <= len(base):
        return base[:column_count]
    return base + [16] * (column_count - len(base))

