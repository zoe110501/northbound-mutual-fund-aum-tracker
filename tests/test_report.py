import json
from pathlib import Path

from openpyxl import load_workbook

from northbound_fund_aum_tracker.report import write_excel_report


def test_write_excel_report_inserts_latest_column_and_marks_changes(tmp_path: Path):
    fund_data = {
        "sheets": {
            "funds": [
                {
                    "sequence": 1,
                    "fund_code": "968000.OF",
                    "name": "示例基金",
                    "english_name": "Example Fund",
                    "inception_date": "2020-01-01",
                    "recognition_approval_date": "2020-01-02",
                    "investment_type": "债券型基金",
                    "investment_region": "中国香港",
                    "recognition_status": "已注册互认",
                    "aum_cny_100m_from_attachment": 10.0,
                    "fund_manager": "张三",
                    "management_company": "示例管理人",
                    "trustee": "示例托管人",
                    "mainland_distributors": "示例代销",
                    "management_fee_pct": 1.0,
                    "management_fee_note": None,
                    "trustee_fee_pct": 0.1,
                    "sales_service_fee_pct": None,
                    "investment_objective": "示例目标",
                    "domicile": "中国香港",
                }
            ],
            "mainland_share_classes": [],
        }
    }
    payload = {
        "run_at_utc": "2026-04-29T00:00:00+00:00",
        "fx_rates_to_usd": {"CNY": 0.14},
        "results": [
            {
                "manager": "示例管理人",
                "global_evidence": [
                    {
                        "target_name": "示例基金",
                        "amount_usd": 154_000_000,
                        "currency": "USD",
                        "source_url": "https://example.com/report.pdf",
                        "context": "Net asset value USD 154 million",
                    }
                ],
                "mainland_evidence": [],
                "errors": [],
            }
        ],
    }
    data_path = tmp_path / "funds.json"
    data_path.write_text(json.dumps(fund_data, ensure_ascii=False), encoding="utf-8")
    output_path = tmp_path / "report.xlsx"

    write_excel_report(payload, data_path, output_path, report_date="2026-04-29")

    workbook = load_workbook(output_path)
    sheet = workbook["合并成基金全称"]
    headers = [cell.value for cell in sheet[1]]
    assert headers[3] == "英文全称"
    latest_col = headers.index("最新基金规模合计(亿人民币）（2026-04-29）") + 1
    assert headers[latest_col - 2] == "基金规模合计(亿人民币)"
    assert sheet.cell(row=2, column=latest_col).value == 11.0
    assert sheet.cell(row=2, column=latest_col).font.color.rgb == "00FF0000"
    assert "https://example.com/report.pdf" in sheet.cell(row=2, column=latest_col).comment.text
